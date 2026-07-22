#!/usr/bin/env python3
"""
Exporta os leads salvos no Postgres (hunter_leads, fonte de verdade desde
2026-07-22 — ver docs/hunter_online_spec.md) para .xlsx — pra abrir direto
no Excel. Sem argumento, exporta todos; com cidade, filtra por ela.

Uso: python exportar_leads_excel.py [cidade]
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from db import SessionLocal
import repository

PASTA_LEADS = Path(__file__).resolve().parent.parent.parent / "leads"

CABECALHO = ["Empresa", "Nicho", "Cidade", "Bairro", "Telefone", "Status", "Google Maps", "Salvo em"]


def exportar(leads: list, destino: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(CABECALHO)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for lead in leads:
        ws.append([
            lead.nome_empresa,
            lead.nicho,
            lead.cidade,
            lead.bairro or "",
            lead.telefone or "a validar",
            lead.status,
            lead.google_maps_url or "",
            lead.created_at.strftime("%d/%m/%Y %H:%M"),
        ])

    for i, titulo in enumerate(CABECALHO, 1):
        largura = max([len(titulo)] + [len(str(row[i - 1].value or "")) for row in ws.iter_rows(min_row=2)])
        ws.column_dimensions[get_column_letter(i)].width = min(largura + 2, 60)

    ws.freeze_panes = "A2"
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destino)
    return destino


def main():
    if SessionLocal is None:
        print("Erro: DATABASE_URL não configurada.")
        sys.exit(1)

    cidade = sys.argv[1] if len(sys.argv) > 1 else None

    db = SessionLocal()
    try:
        leads = repository.listar_leads_hunter(db, cidade=cidade)
    finally:
        db.close()

    if not leads:
        print("Nenhum lead encontrado" + (f" pra '{cidade}'" if cidade else "") + " no banco.")
        sys.exit(1)

    nome_arquivo = f"leads_{cidade}.xlsx".replace(" ", "_").lower() if cidade else "leads_todos.xlsx"
    destino = exportar(leads, PASTA_LEADS / nome_arquivo)
    print(f"Exportado: {destino} ({len(leads)} lead(s))")


if __name__ == "__main__":
    main()
